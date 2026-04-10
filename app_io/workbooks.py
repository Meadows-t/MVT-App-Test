from __future__ import annotations
from pathlib import Path
from datetime import datetime
import pandas as pd

from common.path_utils import ensure_dir
from common.time_utils import build_bins
from unified_inpx_tools import parse_inpx


def indata_dir(root):
    return Path(root) / "_InData"


def timestamped_inputs_path(root):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return indata_dir(root) / ("BaseYear_Observed_Input_" + ts + ".xlsx")


def find_existing_workbooks(root):
    ind = indata_dir(root)
    if not ind.exists():
        return []
    files = list(ind.glob("BaseYear_Observed_Input*.xlsx"))
    files.sort(key=lambda p: p.stat().st_mtime if p.exists() else 0, reverse=True)
    return files


def load_workbook_from_path(xlsx):
    xlsx = Path(xlsx)
    cfg = pd.read_excel(xlsx, sheet_name="Config", engine="openpyxl")
    cfg["ScenarioName"] = cfg["ScenarioName"].astype(str)
    sheets = pd.ExcelFile(xlsx, engine="openpyxl").sheet_names
    classes = pd.read_excel(xlsx, sheet_name="Classes", engine="openpyxl") if "Classes" in sheets else pd.DataFrame()
    return xlsx, cfg, classes


def generate_inputs_workbook(model_folder, project_name, setup_rows, jt_classes, flow_classes, out_path):
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    model_folder = Path(model_folder)
    out_path = Path(out_path)

    inpx_hits = list(model_folder.glob("*.inpx"))
    if not inpx_hits:
        raise FileNotFoundError("No .inpx found in model folder.")
    info = parse_inpx(str(inpx_hits[0]))

    vc_map = {str(x.get("no")).strip(): str(x.get("name")).strip() for x in info.get("vehicle_classes", []) if x.get("no")}
    q_names = [x.get("name") for x in info.get("queue_counters", []) if x.get("name")]
    jt_names = [x.get("name") for x in info.get("jt_measurements", []) if x.get("name")]
    vttm_id_to_name = {str(x.get("no")).strip(): str(x.get("name")).strip() for x in info.get("jt_measurements", []) if x.get("no") and x.get("name")}

    ensure_dir(out_path.parent)
    wb = Workbook()

    ws_cfg = wb.active
    ws_cfg.title = "Config"
    ws_cfg.append(["ScenarioName","PreferredName","Peak","ModelStartHHMM","ModelDurationMin","EvalStartHHMM","EvalDurationMin","ModelPath"])
    for r in setup_rows:
        ws_cfg.append([r["scenario_name"], r.get("preferred_name",""), r.get("peak",""), r["model_start_hhmm"], r["model_duration_min"], r["eval_start_hhmm"], r["eval_duration_min"], str(model_folder)])

    ws_cls = wb.create_sheet("Classes")
    ws_cls.append(["Role","Order","ClassID","ClassName"])
    for i, cid in enumerate(jt_classes[:1], start=1):
        ws_cls.append(["JT", i, str(cid), vc_map.get(str(cid), "")])
    for i, cid in enumerate(flow_classes[:5], start=1):
        ws_cls.append(["Flow", i, str(cid), vc_map.get(str(cid), "")])

    ws_q = wb.create_sheet("Queue")
    ws_q.append(["PreferredName","Peak","ClockTime"] + q_names)
    for r in setup_rows:
        for t in build_bins(str(r["model_start_hhmm"]), int(r["model_duration_min"]), 5):
            ws_q.append([r.get("preferred_name",""), r.get("peak",""), t] + [""]*len(q_names))

    ws_jt = wb.create_sheet("JT")
    palette = ["F0F0F0","E5E5E5","DADADA","CFCFCF","C4C4C4"]
    pal = [PatternFill(start_color=c, end_color=c, fill_type="solid") for c in palette]
    cur = 1
    jt_pairs = [(str(cid), vc_map.get(str(cid),"")) for cid in jt_classes[:1]]

    for i, r in enumerate(setup_rows):
        title = (str(r.get("preferred_name","")) + " – " + str(r.get("peak",""))).strip(" –")
        times = build_bins(str(r["model_start_hhmm"]), int(r["model_duration_min"]), 15)
        header = ["VehicleClassID","VehicleClassName","RouteName"] + times
        fill = pal[i % len(pal)]

        ws_jt.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(header))
        c = ws_jt.cell(row=cur, column=1)
        c.value = title
        c.font = Font(bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = fill
        cur += 1

        for col, h in enumerate(header, start=1):
            cc = ws_jt.cell(row=cur, column=col)
            cc.value = h
            cc.font = Font(bold=True)
            cc.alignment = Alignment(horizontal="center", vertical="center")
            cc.fill = fill

        for cid, cname in jt_pairs:
            for rn in jt_names:
                cur += 1
                ws_jt.cell(row=cur, column=1).value = cid
                ws_jt.cell(row=cur, column=2).value = cname
                ws_jt.cell(row=cur, column=3).value = rn

        cur += 2

    ws_grp = wb.create_sheet("JT_Grouping")
    ws_grp.append(["Group_ID","Segment_Order","Model_Segment_ID","Model_Segment_Name"])
    for mid in sorted(vttm_id_to_name.keys(), key=lambda x: int(x) if str(x).isdigit() else str(x)):
        ws_grp.append(["","", str(mid), vttm_id_to_name.get(str(mid),"")])

    ws_fd = wb.create_sheet("Flow_Definition")
    ws_fd.append(["SiteName","FlowType","Node","FromLink","ToLink","FromArm","ToArm","FlowName"])

    ws_fo = wb.create_sheet("Flow_Observed")
    base_cols = ["SiteName","FromArm","ToArm","FlowType","Node","FromLink","ToLink","FlowName"]
    scenario_names = [str(r.get("scenario_name","")).strip() for r in setup_rows]
    scenario_names = [s for s in scenario_names if s]
    cols = list(base_cols)
    for scen in scenario_names:
        for cid in flow_classes[:5]:
            cols.append(scen + "__Obs_" + str(cid).strip())
    ws_fo.append(cols)

    wb.save(out_path)
